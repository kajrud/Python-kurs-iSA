# week = ['pon', 'wt', 'sr', 'czw', 'pt', 'sob', 'nd']
# weekend = ['sob', 'nd']
# lista = [
#     {"imie": "Kaja",
#      "adresy": ['Adres 1', 'Adres 2']
#      "telefony": {'dom': '456', "praca": '456987'}
#      }
# ]
#
# def wyswietl(text):
#     print(text)
#
# for day in week:
#     if day in weekend:
#         wyswietl('Jest weekend!')

# refaktoryzacja
#
# def fun1():
#     print("fun1")
#
# def fun2():
#     print("fun2")
#
# def fun3():
#     print("fun3")
#
# def bajo():
#     print("Nara, hej")
#     exit()
#
# print("""Hej, multitool!
#       1 - func 1
#       2 - func 2
#       3 - func 3
#       """)
# program = input("Co chcesz zrobić? ")
# funkcje = {"1":fun1,"2": fun2,"3": fun3, "x":bajo}
#
# funkcje[program]() #fun1

# if program == "1":
#     fun1()
# elif program == "2":
#     fun2()
# elif program == "3":
#     fun3()

#openpyxl

import openpyxl
# excel = openpyxl.Workbook()
# excel.save("tabliczka_mnozenia.xlsx")
nazwa_pliku = "tabliczka_mnozenia.xlsx"
excel = openpyxl.load_workbook(nazwa_pliku)
print(excel.sheetnames)
arkusz = excel.active

row = 1
column = 1
max = int(input("Ile chcesz mnożyć? "))
for row in range (1, max + 1):
    for column in range(1, max + 1):
        komorka = arkusz.cell(row, column)
        komorka.value = row * column

excel.save(nazwa_pliku)

# excel = openpyxl.load_workbook('test.xlsx') #nie tworzy nowego pliku, czyta tylko plik
# arkusz = excel.active
# # wyswietla nazwy arkusze = excel.sheetnames
# arkusz = excel['Arkusz1']