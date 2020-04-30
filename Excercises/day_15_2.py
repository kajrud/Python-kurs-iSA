import openpyxl

def dump(file, list):
    excel = openpyxl.load_workbook(file)
    sheet = excel.active

    for row_index, row in enumerate(list):
        for col_index, col in enumerate(row):
            sheet.cell(row_index+1, col_index+1).value = col

    excel.save(file)

def load(file):
    excel = openpyxl.load_workbook(file)
    sheet = excel.active

    data = []
    for row in range(1,sheet.max_row+1):
        cols = []
        for col in range(1, sheet.max_column + 1):
            cols.append(sheet.cell(row, col).value)
        data.append(cols)

    return {"data":data, "rowCount": sheet.max_row, "colCount":sheet.max_column}