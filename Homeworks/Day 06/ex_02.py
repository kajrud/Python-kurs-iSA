# Stwórz program który przyjmie w parametrze ścieżkę do dowolnego pliku (CSV lub Excel - jaki wolicie), który będzie zawierał dane tabelaryczne.
#    W pliku pierwszy wiersz będzie zawierał nazwy kolumn a pozostałe wiersze dane.
#    Ilość kolum i wierszy może być dowolna. Program ma narysować tabelę z danymi, analogicznie do wcześniejszego zadania na rysowanie tabeli.
#    Pamiętajmy by wydzielać części reużywalne do oddzielnych funkcji/modułów (np.: odczyt danych, przygotowanie danych, rysowanie tabeli).
#    Przykład:
#    +------------+------------+------------+
#    | klucz1     | klucz 2    | klucz 3    |
#    +------------+------------+------------+
#    | row 1 col1 | row 1 col2 | row 1 col3 |
#    +------------+------------+------------+
#    | row 2 col1 | row 2 col2 | row 2 col3 |
#    +------------+------------+------------+

import openpyxl
# excel = openpyxl.Workbook()
# excel.save("tabela_danych.xlsx")

nazwa_pliku = "tabela_danych.xlsx"
excel = openpyxl.load_workbook(nazwa_pliku)
arkusz = excel.active
for row in arkusz.iter_rows(min_row=1, max_col=5, max_row=4, values_only=True):
    num = len(row)
    upndown = "+" + "------------+" * num
    print(upndown)
    counter = 0
    for column in row:
        data = (row[counter])
        data = str(data)
        if len(data) <= 7:
            print("|{:11.8s}".format(data) + " ", end="")
            counter += 1
        else:
            print("|{:9.7s}".format(data) + "...", end="")
            counter += 1
    print("|")
print(upndown)






